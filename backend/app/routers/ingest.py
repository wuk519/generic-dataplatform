import csv
import gzip
import io
import json
from datetime import datetime, timezone
from pathlib import PurePosixPath

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    UploadFile,
    status,
)
from sqlalchemy.ext.asyncio import AsyncSession

from ..db import get_db
from ..deps import (
    Principal,
    acting_user_id,
    get_principal,
    is_admin_principal,
)
from ..ingest_core import (
    BATCH_SIZE,
    DROP,
    infer_scalar,
    insert_batch,
    normalize_record,
    set_source_descriptions,
    sources_owned_by_others,
)
from ..models import ApiKey
from ..schemas import IngestRecord, IngestResponse

router = APIRouter(prefix="/ingest", tags=["ingest"])

# Extensions we recognize for format auto-detection.
_EXT_FORMATS = {
    ".csv": "csv",
    ".tsv": "tsv",
    ".tab": "tsv",
    ".ndjson": "ndjson",
    ".jsonl": "ndjson",
    ".json": "json",
    ".xlsx": "xlsx",
    ".xlsm": "xlsx",
}


def _touch_api_key(principal: Principal) -> None:
    if isinstance(principal, ApiKey):
        principal.last_used_at = datetime.now(timezone.utc)


def _infer_format(filename: str | None, peek: bytes) -> str | None:
    """Infer the upload format from filename extension first, then content."""
    if filename:
        ext = PurePosixPath(filename).suffix.lower()
        if ext in _EXT_FORMATS:
            return _EXT_FORMATS[ext]

    head = peek.lstrip()
    if head.startswith(b"["):
        return "json"
    if head.startswith(b"{"):
        return "ndjson"
    first_line = head.split(b"\n", 1)[0]
    if b"\t" in first_line:
        return "tsv"
    if b"," in first_line:
        return "csv"
    return None


def _typed_row(row: dict) -> dict:
    """Apply scalar type-inference to a CSV/TSV/Excel row, dropping empty cells."""
    out: dict = {}
    for k, raw in row.items():
        if k is None or k == "":
            continue
        val = infer_scalar(raw)
        if val is DROP:
            continue
        out[str(k)] = val
    return out


def _parse_delimited(text: str, delimiter: str | None) -> list[dict]:
    """Parse CSV/TSV text into typed dict rows. Sniffs the delimiter if unknown."""
    sample = text[:8192]
    if delimiter is None:
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
        except csv.Error:
            delimiter = ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [_typed_row(row) for row in reader]


def _parse_xlsx(data: bytes) -> list[dict]:
    """Parse the first worksheet of an .xlsx workbook into typed dict rows.

    Row 1 is treated as the header. openpyxl is imported lazily so the rest of
    the upload path works even if the optional dependency is missing.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover - dependency missing
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Excel support requires the 'openpyxl' package. Re-run the install "
            "script (pip install -e .) to add it.",
        ) from e

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    headers = [str(h) if h is not None else "" for h in header]

    out: list[dict] = []
    for raw_row in rows:
        record: dict = {}
        for key, cell in zip(headers, raw_row):
            if not key:
                continue
            val = infer_scalar(cell)
            if val is DROP:
                continue
            record[key] = val
        if record:
            out.append(record)
    wb.close()
    return out


@router.post("", response_model=IngestResponse)
async def ingest(
    body: list[IngestRecord] | IngestRecord,
    description: str | None = Query(
        None, description="Optional description to set on the source(s) ingested"
    ),
    principal: Principal = Depends(get_principal),
    db: AsyncSession = Depends(get_db),
) -> IngestResponse:
    items = body if isinstance(body, list) else [body]
    try:
        rows = [normalize_record(r.model_dump()) for r in items]
    except ValueError as e:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(e)) from e

    owner = acting_user_id(principal)
    touched = {r["source_id"] for r in rows}
    if not is_admin_principal(principal):
        conflicts = await sources_owned_by_others(db, touched, owner)
        if conflicts:
            raise HTTPException(
                status.HTTP_403_FORBIDDEN,
                f"Source(s) owned by another user: {', '.join(sorted(conflicts))}",
            )

    for i in range(0, len(rows), BATCH_SIZE):
        await insert_batch(db, rows[i : i + BATCH_SIZE], owner_id=owner)

    if description is not None:
        await set_source_descriptions(db, touched, description)

    _touch_api_key(principal)
    await db.commit()
    return IngestResponse(accepted=len(rows))


@router.post("/upload", response_model=IngestResponse)
async def upload(
    file: UploadFile = File(...),
    source_id: str | None = Form(None),
    format: str | None = Form(None),
    description: str | None = Form(None),
    principal: Principal = Depends(get_principal),
    db: AsyncSession = Depends(get_db),
) -> IngestResponse:
    data = await file.read()
    filename = file.filename or ""

    # Transparently decompress .gz, then re-infer from the inner extension.
    if filename.lower().endswith(".gz") or data[:2] == b"\x1f\x8b":
        try:
            data = gzip.decompress(data)
        except OSError as e:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST, f"Could not decompress gzip: {e}"
            ) from e
        if filename.lower().endswith(".gz"):
            filename = filename[:-3]

    requested = (format or "").strip().lower()
    if requested in ("", "auto"):
        fmt = _infer_format(filename, data[:4096])
        if not fmt:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                "Could not infer file format. Use a recognized extension "
                "(.csv, .tsv, .ndjson, .json, .xlsx; optionally .gz) or pass "
                "format=csv|tsv|ndjson|json|xlsx.",
            )
    else:
        fmt = requested

    accepted = 0
    batch: list[dict] = []
    touched: set[str] = set()
    owner = acting_user_id(principal)
    is_admin = is_admin_principal(principal)

    async def flush() -> None:
        nonlocal accepted, batch
        if batch:
            if not is_admin:
                conflicts = await sources_owned_by_others(
                    db, {r["source_id"] for r in batch}, owner
                )
                if conflicts:
                    raise HTTPException(
                        status.HTTP_403_FORBIDDEN,
                        f"Source(s) owned by another user: {', '.join(sorted(conflicts))}",
                    )
            await insert_batch(db, batch, owner_id=owner)
            await db.commit()
            accepted += len(batch)
            batch = []

    def add(obj: dict) -> None:
        row = normalize_record(obj, source_id)
        touched.add(row["source_id"])
        batch.append(row)

    try:
        if fmt == "ndjson":
            for raw_line in data.splitlines():
                line = raw_line.strip()
                if not line:
                    continue
                add(json.loads(line))
                if len(batch) >= BATCH_SIZE:
                    await flush()
            await flush()

        elif fmt == "json":
            arr = json.loads(data)
            if not isinstance(arr, list):
                raise HTTPException(
                    status.HTTP_400_BAD_REQUEST, "JSON body must be an array"
                )
            for obj in arr:
                add(obj)
                if len(batch) >= BATCH_SIZE:
                    await flush()
            await flush()

        elif fmt in ("csv", "tsv"):
            text = data.decode("utf-8-sig")
            delimiter = "\t" if fmt == "tsv" else None
            for record in _parse_delimited(text, delimiter):
                add(record)
                if len(batch) >= BATCH_SIZE:
                    await flush()
            await flush()

        elif fmt == "xlsx":
            for record in _parse_xlsx(data):
                add(record)
                if len(batch) >= BATCH_SIZE:
                    await flush()
            await flush()

        else:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Unknown format: {fmt}")

    except json.JSONDecodeError as e:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST, f"Invalid JSON (format={fmt}): {e}"
        ) from e
    except ValueError as e:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(e)) from e

    if description is not None and description.strip():
        await set_source_descriptions(db, touched, description.strip())

    _touch_api_key(principal)
    await db.commit()
    return IngestResponse(accepted=accepted, format=fmt)
